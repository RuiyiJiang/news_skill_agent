from __future__ import annotations

from datetime import datetime

from openpyxl import load_workbook

from app.excel_writer import write_news_report
from app.models import NewsItem


def test_write_news_report_with_items(tmp_path) -> None:
    item = NewsItem(
        title="Test News",
        summary="Summary",
        published_at=datetime.fromisoformat("2026-04-01T09:00:00+08:00"),
        url="https://example.com/article",
        source="Example",
        label="技术创新",
        collected_at=datetime.fromisoformat("2026-04-01T09:00:00+08:00"),
        date_parse_status="parsed",
    )
    output = write_news_report([item], tmp_path, item.collected_at, report_name="news_report_all")
    workbook = load_workbook(output)
    sheet = workbook["news"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == ["Source", "Title", "Summary", "Published At", "URL", "标签"]
    assert output.name == "news_report_all_2026-04-01_090000.xlsx"
    assert sheet["B2"].value == "Test News"
    assert sheet["D2"].value == "2026-04-01 09:00:00"
    assert sheet["F2"].value == "技术创新"


def test_write_news_report_with_empty_items(tmp_path) -> None:
    now = datetime.fromisoformat("2026-04-01T09:00:00+08:00")
    output = write_news_report([], tmp_path, now, report_name="news_report_filtered")
    workbook = load_workbook(output)
    sheet = workbook["news"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == ["Source", "Title", "Summary", "Published At", "URL", "标签"]
    assert output.name == "news_report_filtered_2026-04-01_090000.xlsx"
    assert sheet.max_row == 1


def test_write_news_report_with_extra_status_rows(tmp_path) -> None:
    now = datetime.fromisoformat("2026-04-01T09:00:00+08:00")
    output = write_news_report(
        [],
        tmp_path,
        now,
        report_name="news_report_filtered",
        extra_rows=[
            {
                "Source": "联合利华",
                "Title": "本次未抓到符合时间窗口的资讯",
                "Summary": "抓取成功，但当前时间窗口内没有可入表内容。",
                "Published At": "",
                "URL": "https://www.unilever.com/news/news-search/",
                "标签": "状态记录",
            }
        ],
    )
    workbook = load_workbook(output)
    sheet = workbook["news"]
    assert sheet.max_row == 2
    assert sheet["A2"].value == "联合利华"
    assert sheet["B2"].value == "本次未抓到符合时间窗口的资讯"
    assert sheet["F2"].value == "状态记录"
